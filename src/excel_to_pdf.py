

import os
import math
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, A3, landscape, portrait
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.units import inch
import pyexcel as p


def convert_xls_to_xlsx(xls_path, xlsx_path=None):
    """Convert the old .xls file to .xlsx format"""
    if xlsx_path is None:
        xlsx_path = os.path.splitext(xls_path)[0] + '.xlsx'
    p.save_book_as(file_name=xls_path, dest_file_name=xlsx_path)
    return xlsx_path


def determine_page_format(num_columns, max_column_width=None):
    """
    Determine the optimal page size and orientation based on table dimensions.
    
    Args:
        num_columns (int): Number of columns in the table.
        max_column_width (float, optional): Maximum column width if available.
        
    Returns:
        tuple: (pagesize, orientation function)
    """
    # Define thresholds for decision making
    if num_columns <= 5:
        # Few columns, likely to fit on portrait A4
        return A4, portrait
    elif num_columns <= 8:
        # Medium number of columns, use landscape A4
        return A4, landscape
    elif num_columns <= 12:
        # Many columns, use portrait A3
        return A3, portrait
    else:
        # Lots of columns, use landscape A3
        return A3, landscape


def is_effectively_empty(value):
    """
    Return True if the cell value is considered empty.
    
    Empty means:
      - The value is None.
      - The value is a float and math.isnan(value) is True.
      - The value is a string that is empty (after stripping whitespace).
    """
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def excel_to_pdf(excel_path, pdf_path=None, sheet_name=None, max_rows_per_table=50):
    """
    Convert Excel file to PDF with adaptive page size based on content,
    removing columns that contain only NaN (or empty) values.
    
    Args:
        excel_path (str): Path to the Excel file.
        pdf_path (str, optional): Path for the output PDF file.
        sheet_name (str, optional): Name of the sheet to convert.
        max_rows_per_table (int): Maximum rows per table before splitting.
        
    Returns:
        str: Path to the created PDF file.
    """
    if excel_path.endswith('.xls'):
        excel_path = convert_xls_to_xlsx(excel_path)

    if pdf_path is None:
        pdf_path = os.path.splitext(excel_path)[0] + '.pdf'
    
    # Load Excel file
    wb = load_workbook(excel_path)
    sheets = [sheet_name] if sheet_name else wb.sheetnames

    # Create paragraph styles for cell content
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        name='HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.white,
        leading=12
    )
    cell_style = ParagraphStyle(
        name='CellStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        alignment=TA_LEFT,
        leading=10  # Line spacing
    )
    
    elements = []
    
    # Determine the effective maximum number of columns among all sheets (after filtering out empty ones)
    global_effective_max_columns = 0
    for sh in sheets:
        sheet = wb[sh]
        effective_cols = 0
        for col in range(1, sheet.max_column + 1):
            # Check if any cell in the column is non-empty
            for row in range(1, sheet.max_row + 1):
                if not is_effectively_empty(sheet.cell(row=row, column=col).value):
                    effective_cols += 1
                    break
        global_effective_max_columns = max(global_effective_max_columns, effective_cols)
    
    # Determine optimal page format based on effective column count
    pagesize, orientation_func = determine_page_format(global_effective_max_columns)
    
    # Create the document with determined format
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=orientation_func(pagesize),
        leftMargin=10,
        rightMargin=10,
        topMargin=15,
        bottomMargin=15
    )
    
    # Process each sheet
    for sheet_idx, current_sheet in enumerate(sheets):
        sheet = wb[current_sheet]
        
        # Determine which columns to keep (those with at least one non-empty cell)
        columns_to_keep = []
        for col in range(1, sheet.max_column + 1):
            for row in range(1, sheet.max_row + 1):
                if not is_effectively_empty(sheet.cell(row=row, column=col).value):
                    columns_to_keep.append(col)
                    break
        
        # If no columns have valid data, skip this sheet.
        if not columns_to_keep:
            continue
        
        # Calculate appropriate column widths (only for kept columns)
        max_col_width = 130  # Maximum column width in points
        min_col_width = 40   # Minimum column width in points
        if pagesize == A3:
            max_col_width = 150  # Allow wider columns on A3
        
        col_widths = []
        for col in columns_to_keep:
            max_length = 0
            # Sample first 100 rows for efficiency
            for row in range(1, min(100, sheet.max_row) + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    content_length = len(str(cell.value))
                    # Cap the length for width calculation at 30 characters
                    max_length = max(max_length, min(content_length, 30))
            # Adjust multiplier based on page format (narrower columns for A4, wider for A3)
            multiplier = 5.5 if pagesize == A4 else 6.0
            width = min(max(min_col_width, max_length * multiplier), max_col_width)
            col_widths.append(width)
        
        # Build the header row from the kept columns
        header_row = []
        # Using row 1 as header (or adjust if your header is in another row)
        for col in columns_to_keep:
            cell_value = sheet.cell(row=1, column=col).value
            header_row.append(Paragraph(str(cell_value or ""), header_style))
        
        # Process data rows in chunks to avoid huge tables that might get chopped
        row_count = sheet.max_row
        # Start after header row
        start_row = 2  
        while start_row <= row_count:
            end_row = min(start_row + max_rows_per_table - 1, row_count)
            
            # Create data for this chunk, starting with the header row
            chunk_data = [header_row]
            for row_idx in range(start_row, end_row + 1):
                data_row = []
                for col in columns_to_keep:
                    cell = sheet.cell(row=row_idx, column=col)
                    cell_value = cell.value or ""
                    data_row.append(Paragraph(str(cell_value), cell_style))
                chunk_data.append(data_row)
            
            # Create table for this chunk
            table = Table(chunk_data, colWidths=col_widths, repeatRows=1)
            
            # Style the table
            table_style = TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                
                # Row background colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                
                # Cell padding
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3)
            ])
            
            table.setStyle(table_style)
            table.hAlign = 'LEFT'
            table.spaceBefore = 5
            table.spaceAfter = 15
            
            elements.append(table)
            
            # Uncomment below if you wish to add a continuation note when splitting tables
            # if end_row < row_count:
            #     continuation = Paragraph(f"Table continues... (Rows {start_row}-{end_row} of {row_count})", styles['Italic'])
            #     elements.append(continuation)
            #     elements.append(Spacer(1, 0.2 * inch))
            
            start_row = end_row + 1
        
        # Add page break between sheets (except for the last sheet)
        if sheet_idx < len(sheets) - 1:
            elements.append(PageBreak())
    
    # Build PDF
    doc.build(elements)
    
    return pdf_path